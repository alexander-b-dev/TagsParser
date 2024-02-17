import json
import os
from kivy.config import Config
from kivy.core.window import Window
from kivy.uix.behaviors import ButtonBehavior
from kivymd.app import MDApp
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.dialog import MDDialog
from kivymd.uix.label import MDLabel
from kivymd.uix.button import MDRaisedButton, MDIconButton, MDFlatButton
from kivymd.uix.list import MDList, OneLineListItem
from kivymd.uix.menu import MDDropdownMenu
from kivymd.uix.relativelayout import MDRelativeLayout
from kivymd.uix.scrollview import MDScrollView
from kivymd.uix.stacklayout import MDStackLayout
from kivymd.uix.textfield import MDTextField
from openpyxl.reader.excel import load_workbook


class PressableOneLineItem(OneLineListItem, ButtonBehavior):
    pass


class DialogContent(MDBoxLayout):

    def __init__(self, *args, **kwargs):
        inputData = kwargs.pop("dataDict")
        self.columnName = ""
        self.creationName = ""
        self.defaultValue = ""
        self.multipleValue = ""
        self.tagsSet = set()

        super().__init__(*args, **kwargs)
        self.orientation = "vertical"

        self.columnNameTextField = MDTextField(hint_text="Название колонки")
        self.defaultTextField = MDTextField(hint_text="Значение, если ничего не нашлось")
        self.multipleTextField = MDTextField(hint_text="Значение, если нашлось слишком много")
        self.tagsList = MDList()

        if inputData:
            self.columnNameTextField.text = self.columnName = inputData["columnName"]
            self.creationName = inputData["columnName"]
            self.defaultTextField.text = self.defaultValue = inputData["default"]
            self.multipleTextField.text = self.multipleValue = inputData["multiple"]
            for tag in sorted(list(inputData["all"]), key=str.casefold):
                self.tagsSet.add(tag)
                self.tagsList.add_widget(PressableOneLineItem(text=tag,
                                                              on_release=self.askDelete))

        self.columnNameTextField.bind(text=self.textValueChange)
        self.defaultTextField.bind(text=self.textValueChange)
        self.multipleTextField.bind(text=self.textValueChange)

        self.add_widget(self.columnNameTextField)
        self.add_widget(self.defaultTextField)
        self.add_widget(self.multipleTextField)

        scroll = MDScrollView()
        scroll.size_hint = (1, 1)
        scroll.add_widget(self.tagsList)
        self.add_widget(scroll)

    def textValueChange(self, field, val):
        if field == self.defaultTextField:
            self.defaultValue = val
        elif field == self.columnNameTextField:
            self.columnName = val
        elif field == self.multipleTextField:
            self.multipleValue = val

    def askDelete(self, tag):
        print(tag)
        confirmDialog = MDDialog(
            text="Удалить тег \"%s\" ?" % tag.text,
            buttons=[
                MDFlatButton(
                    text="Отмена",
                    theme_text_color="Custom",
                    text_color=self.theme_cls.primary_color,
                    on_press=lambda x: confirmDialog.dismiss()
                ),
                MDFlatButton(
                    text="OK",
                    theme_text_color="Custom",
                    text_color=self.theme_cls.primary_color,
                    on_press=lambda x: self.removeTag(confirmDialog, tag.text)
                )
            ],
        )
        confirmDialog.open()
        return True

    def removeTag(self, dialog, tag):
        self.tagsSet.remove(tag)
        tagToRemove = None
        for tagElement in self.tagsList.children:
            if tagElement.text == tag:
                tagToRemove = tagElement
                break
        if tagToRemove:
            self.tagsList.remove_widget(tagToRemove)
        dialog.dismiss()


class MainApp(MDApp):

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        self.newTags = None
        self.selectedTag = None
        self.confirmDialog = None
        self.errDialog = None
        self.dialog = None
        self.usedTagsContainer = None
        self.unusedTagsContainer = None
        self.columnsMenu = None
        self.columnSelector = None
        self.mainContainer = None
        self.filePath = None
        self.xlFile = None

        self.theme_cls.theme_style = "Dark"
        self.theme_cls.primary_palette = "Lime"

        self.mainWidget = MDRelativeLayout()
        self.mainWidget.size_hint = (1, 1)
        self.mainWidget.add_widget(MDLabel(text="Положи сюда файл", halign="center"))

        if os.path.exists("config.json"):
            with open("config.json", "r") as cfg:
                self.conf = json.loads(cfg.read())
                self.conf["unused"] = set(self.conf["unused"])
                for col in self.conf["used"].values():
                    col["all"] = set(col["all"])
        else:
            self.conf = {"unused": set(),
                         "used": {}}  # {columnName: {default: value,
            #                                         multiple: value,
            #                                         all: []}}

    def prepareWorkspace(self):
        self.mainWidget.clear_widgets()
        controlArea = MDBoxLayout(orientation="horizontal")
        controlArea.size_hint = (1, 0.1)
        controlArea.radius = (25, 25, 25, 25)
        controlArea.padding = 10
        controlArea.spacing = 5
        controlArea.md_bg_color = self.theme_cls.bg_dark
        doneBtn = MDRaisedButton(text="Сохранить",
                                 on_release=lambda x: self.saveResults())
        doneBtn.pos_hint = {"center_y": 0.5}
        fileNameLbl = MDLabel(text=self.filePath.split("\\")[-1])
        fileNameLbl.pos_hint = {"center_y": 0.5}
        fileNameLbl.bind(size=fileNameLbl.setter("texture_size"))
        self.columnSelector = MDRaisedButton(text="Выбери колонку с тегами",
                                             on_release=lambda x: self.openColumnsMenu())
        self.columnSelector.pos_hint = {"center_y": 0.5,
                                        "right": 1}
        menuItems = [
            {
                "text": chr(col + 65),
                "on_release": lambda x=(chr(col + 65)): self.columnSelected(x),
            } for col in range(0, self.xlFile.worksheets[0].max_column)
        ]
        self.columnsMenu = MDDropdownMenu(caller=self.columnSelector, items=menuItems)

        controlArea.add_widget(fileNameLbl)
        controlArea.add_widget(self.columnSelector)
        controlArea.add_widget(doneBtn)

        self.mainContainer = MDStackLayout()
        self.mainContainer.size_hint = (1, 1)
        self.mainContainer.padding = 10
        self.mainContainer.spacing = 10
        self.mainContainer.orientation = "tb-rl"
        self.mainContainer.add_widget(controlArea)
        self.mainWidget.add_widget(self.mainContainer)

    def openColumnsMenu(self):
        self.columnsMenu.open()

    def columnSelected(self, selectedColumn):
        self.columnSelector.text = selectedColumn
        self.columnsMenu.dismiss()
        listsContainer = MDBoxLayout(orientation="horizontal")
        listsContainer.size_hint = (1, 0.9)
        listsContainer.radius = (25, 25, 25, 25)
        listsContainer.padding = 10
        listsContainer.spacing = 50
        listsContainer.md_bg_color = self.theme_cls.bg_dark

        allTags = []
        xlSheet = self.xlFile.worksheets[0]
        for tagCell in range(2, xlSheet.max_row):
            cellVal = xlSheet["%s%s" % (selectedColumn, tagCell)].value
            if cellVal:
                allTags.extend(map(str.strip, cellVal.replace("\n", " ").split(",")))

        allSheetTags = set(allTags)
        if "" in allSheetTags:
            allSheetTags.remove("")
        allTags = self.getAlltagsSet()
        self.unusedTagsContainer = MDList()
        self.newTags = list(allSheetTags.difference(allTags))
        self.fillUnusedContainer(self.newTags,
                                 list(allSheetTags.intersection(self.conf["unused"])))
        unusedTagsScroll = MDScrollView()
        unusedTagsScroll.scroll_type = ['bars']
        unusedTagsScroll.bar_color = self.theme_cls.primary_color
        unusedTagsScroll.bar_inactive_color = self.theme_cls.accent_color
        unusedTagsScroll.bar_width = 5
        unusedTagsScroll.add_widget(self.unusedTagsContainer)
        listsContainer.add_widget(unusedTagsScroll)

        structureContainer = MDRelativeLayout()
        self.usedTagsContainer = MDList()
        configuredColumns = sorted(list(self.conf["used"].keys()), key=str.casefold)
        for col in configuredColumns:
            lstItem = PressableOneLineItem(text=col,
                                           on_press=lambda x: self.columnClicked(col))
            self.usedTagsContainer.add_widget(lstItem)
        addBtn = MDIconButton(icon="plus",
                              theme_icon_color="Custom",
                              icon_color=self.theme_cls.bg_darkest,
                              on_press=lambda x: self.openConfigureColumnPopup(""))
        addBtn.pos_hint = {"center_x": 0.9,
                           "center_y": 0.1}
        addBtn.md_bg_color = self.theme_cls.primary_color
        usedTagsContainerScroll = MDScrollView()
        usedTagsContainerScroll.scroll_type = ['bars']
        usedTagsContainerScroll.bar_color = self.theme_cls.primary_color
        usedTagsContainerScroll.bar_inactive_color = self.theme_cls.accent_color
        usedTagsContainerScroll.bar_width = 5
        usedTagsContainerScroll.add_widget(self.usedTagsContainer)
        structureContainer.add_widget(usedTagsContainerScroll)
        structureContainer.add_widget(addBtn)
        listsContainer.add_widget(structureContainer)

        self.mainContainer.add_widget(listsContainer)

    def fillUnusedContainer(self, newTags, unusedTags):
        self.unusedTagsContainer.clear_widgets()

        for newTag in sorted(newTags, key=str.casefold):
            lstItem = PressableOneLineItem(text=newTag,
                                           on_release=self.selectTag)
            lstItem.divider_color = self.theme_cls.primary_color
            lstItem.theme_text_color = "Custom"
            lstItem.text_color = self.theme_cls.primary_color
            self.unusedTagsContainer.add_widget(lstItem)

        for knownTag in sorted(unusedTags, key=str.casefold):
            lstItem = PressableOneLineItem(text=knownTag)
            self.unusedTagsContainer.add_widget(lstItem)

    def getAlltagsSet(self):
        tagsSet = set()
        tagsSet.update(self.conf["unused"])
        for colData in self.conf["used"].values():
            tagsSet.update(colData["all"])
        return tagsSet

    def openConfigureColumnPopup(self, colName):
        if colName != "":
            data = dict()
            data.update(self.conf["used"][colName])
            data["columnName"] = colName
        else:
            data = None

        self.dialog = MDDialog(
            type="custom",
            size_hint=(1, 1),
            content_cls=DialogContent(dataDict=data,
                                      size_hint=(None, None),
                                      width=self.mainWidget.width * 0.6,
                                      height=self.mainWidget.height * 0.8),
            buttons=[
                MDFlatButton(
                    text="Удалить столбец",
                    theme_text_color="Custom",
                    text_color=self.theme_cls.primary_color,
                    on_press=lambda x: self.showConfirmDialog("Удалить столбец?", lambda: self.delColumn(colName)),
                    disabled=(colName == "")
                ),
                MDFlatButton(
                    text="Отмена",
                    theme_text_color="Custom",
                    text_color=self.theme_cls.primary_color,
                    on_press=lambda x: self.dialog.dismiss(),
                ),
                MDFlatButton(
                    text="OK",
                    theme_text_color="Custom",
                    text_color=self.theme_cls.primary_color,
                    on_press=lambda x: self.addOrModifyColumn(self.dialog.content_cls.columnName,
                                                              self.dialog.content_cls.creationName,
                                                              self.dialog.content_cls.defaultValue,
                                                              self.dialog.content_cls.multipleValue,
                                                              self.dialog.content_cls.tagsSet),
                ),
            ],
        )
        self.dialog.open()
        return True

    def addOrModifyColumn(self, columnName, creationName, defaultValue, multipleValue, tagsSet):
        if columnName == "":
            self.showErrDialog("Надо назвать колонку")
            return
        if columnName != creationName:
            if columnName in self.conf["used"]:
                self.showErrDialog("Такая колонка уже есть")
                return
            itemToDelete = None
            for item in self.usedTagsContainer.children:
                if item.text == creationName:
                    itemToDelete = item
                    break
            if itemToDelete:
                self.usedTagsContainer.remove_widget(itemToDelete)
                del self.conf["used"][creationName]

        if columnName not in self.conf["used"]:
            self.usedTagsContainer.add_widget(
                PressableOneLineItem(text=columnName,
                                     on_press=lambda x: self.columnClicked(columnName)
                                     )
            )
        colData = self.conf["used"].setdefault(columnName, dict())
        colData["default"] = defaultValue
        colData["multiple"] = multipleValue
        if "all" not in colData.keys():
            colData["all"] = set()
        elif colData["all"] != tagsSet:
            newNewTagsList = [tagEl.text for tagEl in self.unusedTagsContainer.children
                              if tagEl.text_color == self.theme_cls.primary_color]
            newKnownTagsList = [tagEl.text for tagEl in self.unusedTagsContainer.children
                                if tagEl.text_color != self.theme_cls.primary_color]
            for deletedTag in colData["all"].difference(tagsSet):
                if deletedTag in self.newTags:
                    newNewTagsList.append(deletedTag)
                else:
                    newKnownTagsList.append(deletedTag)
                    self.conf["unused"].add(deletedTag)

            self.fillUnusedContainer(newNewTagsList, newKnownTagsList)
            colData["all"] = tagsSet
        self.dialog.dismiss()

    def columnClicked(self, colName):
        if not self.selectedTag:
            self.openConfigureColumnPopup(colName)
        else:
            self.dialog = MDDialog(
                text="Добавить тег \"%s\" к столбцу \"%s\"?" % (self.selectedTag, colName),
                buttons=[
                    MDFlatButton(
                        text="Отмена",
                        theme_text_color="Custom",
                        text_color=self.theme_cls.primary_color,
                        on_press=lambda x: self.dialog.dismiss(),
                    ),
                    MDFlatButton(
                        text="OK",
                        theme_text_color="Custom",
                        text_color=self.theme_cls.primary_color,
                        on_press=lambda x: self.addTagToColumn(colName),
                    ),
                ],
            )
            self.dialog.open()
        return True

    def delColumn(self, colName):
        colData = self.conf["used"][colName]
        newNewTagsList = [tagEl.text for tagEl in self.unusedTagsContainer.children
                          if tagEl.text_color == self.theme_cls.primary_color]
        newKnownTagsList = [tagEl.text for tagEl in self.unusedTagsContainer.children
                            if tagEl.text_color != self.theme_cls.primary_color]
        for deletedTag in colData["all"]:
            if deletedTag in self.newTags:
                newNewTagsList.append(deletedTag)
            else:
                newKnownTagsList.append(deletedTag)
                self.conf["unused"].add(deletedTag)

        self.fillUnusedContainer(newNewTagsList, newKnownTagsList)

        colElToDelete = next(colEl for colEl in self.usedTagsContainer.children if colEl.text == colName)
        self.usedTagsContainer.remove_widget(colElToDelete)
        del self.conf["used"][colName]

        self.confirmDialog.dismiss()
        self.dialog.dismiss()

    def showConfirmDialog(self, txt, callback):
        self.confirmDialog = MDDialog(
            text=txt,
            buttons=[
                MDFlatButton(
                    text="Отмена",
                    theme_text_color="Custom",
                    text_color=self.theme_cls.primary_color,
                    on_press=lambda x: self.confirmDialog.dismiss(),
                ),
                MDFlatButton(
                    text="OK",
                    on_press=lambda x: callback()
                )
            ],
        )
        self.confirmDialog.open()

    def showErrDialog(self, txt):
        self.errDialog = MDDialog(
            text=txt,
            buttons=[
                MDFlatButton(
                    text="Ok",
                    on_press=lambda x: self.errDialog.dismiss()
                )
            ],
        )
        self.errDialog.open()

    def selectTag(self, tagEl: PressableOneLineItem):
        if self.selectedTag:
            prevTag: PressableOneLineItem = next(tag for tag in self.unusedTagsContainer.children
                                                 if tag.text == self.selectedTag)
            if prevTag.text in self.newTags:
                prevTag.theme_text_color = "Custom"
                prevTag.text_color = self.theme_cls.primary_color
                prevTag.bg_color = self.theme_cls.bg_dark
                prevTag.divider_color = self.theme_cls.primary_color
            else:
                prevTag.theme_text_color = "Custom"
                prevTag.text_color = self.theme_cls.text_color
                prevTag.bg_color = self.theme_cls.bg_dark
                prevTag.divider_color = self.theme_cls.bg_light
            if self.selectedTag == tagEl.text:
                self.selectedTag = None
                return True
        self.selectedTag = tagEl.text
        tagEl.theme_text_color = "Custom"
        tagEl.text_color = self.theme_cls.bg_dark
        tagEl.bg_color = self.theme_cls.primary_color
        tagEl.divider_color = self.theme_cls.bg_light
        return True

    def addTagToColumn(self, colName):
        tagEl: PressableOneLineItem = next(tag for tag in self.unusedTagsContainer.children
                                           if tag.text == self.selectedTag)
        self.unusedTagsContainer.remove_widget(tagEl)
        self.conf["used"][colName]["all"].add(self.selectedTag)
        if self.selectedTag not in self.newTags:
            self.conf["unused"].remove(self.selectedTag)
        self.selectedTag = None
        self.dialog.dismiss()
        return True

    def loadFile(self, filePath: str):
        if filePath[-5:] == ".xlsx":
            self.xlFile = load_workbook(filename=filePath)
            self.filePath = filePath
            self.prepareWorkspace()

    def saveResults(self):
        xlSheet = self.xlFile.worksheets[0]
        selectedColumn = ord(self.columnSelector.text) - 64
        for num, col in enumerate(self.conf["used"].keys()):
            xlSheet.insert_cols(selectedColumn + num + 1)
            xlSheet.cell(row=1, column=selectedColumn + num + 1, value=col)

        for row in range(2, xlSheet.max_row):
            origCellData = xlSheet.cell(row=row, column=selectedColumn).value
            if not origCellData:
                continue
            origCellData = set(map(str.strip, origCellData.replace("\n", " ").split(",")))

            for num, colData in enumerate(self.conf["used"].values()):
                cell = xlSheet.cell(row=row, column=selectedColumn + num + 1)
                intersection = colData["all"].intersection(origCellData)
                if len(intersection) == 0:
                    cell.value = colData["default"]
                elif len(intersection) == 1:
                    cell.value = next(iter(intersection))
                else:
                    if colData["multiple"] == "[all]":
                        cell.value = ", ".join(intersection)
                    else:
                        cell.value = colData["multiple"]
        self.xlFile.save(self.filePath.replace(".xlsx", "_updated.xlsx"))

        with open("config.json", "w") as cfg:
            self.conf["unused"].update(self.newTags)
            config = {"unused": list(self.conf["unused"]), "used": dict()}
            for col, colData in self.conf["used"].items():
                config["used"][col] = {"default": colData["default"],
                                       "multiple": colData["multiple"],
                                       "all": list(colData["all"])}
            cfg.write(json.dumps(config))

        self.dialog = MDDialog(
            text="Готово!",
            buttons=[
                MDFlatButton(
                    text="OK",
                    theme_text_color="Custom",
                    text_color=self.theme_cls.primary_color,
                    on_press=lambda x: self.dialog.dismiss(),
                ),
            ],
        )
        self.dialog.open()

    def build(self):
        Window.bind(on_drop_file=lambda window, file, x, y: self.loadFile(file.decode("utf-8")))
        return self.mainWidget


if __name__ == '__main__':
    Config.set('input', 'mouse', 'mouse,multitouch_on_demand')
    abspath = os.path.abspath(__file__)
    dname = os.path.dirname(abspath)
    os.chdir(dname)
    MainApp().run()
